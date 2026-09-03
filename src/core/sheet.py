# -*- coding: utf-8 -*-
"""Excel 工作簿操作 — 多工作表合并(F01) / 单簿内表合并(F05) / 工作簿拆分CSV(F09) / 超宽列拆分(F10)"""
import os
import csv
import datetime
import contextlib
from pathlib import Path

try:
    import openpyxl
    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False

try:
    import xlrd
    HAVE_XLRD = True
except ImportError:
    xlrd = None
    HAVE_XLRD = False


def _cell_to_str(c):
    """将单元格值转为字符串，处理 datetime/float 特殊格式（#18）"""
    if c is None:
        return ""
    if isinstance(c, datetime.datetime):
        return c.strftime("%Y-%m-%d") + (c.strftime(" %H:%M:%S") if (c.hour or c.minute or c.second) else "")
    if isinstance(c, datetime.date):
        return c.strftime("%Y-%m-%d")
    if isinstance(c, float):
        # 避免科学计数和精度尾巴
        return "%g" % c
    return str(c)


def _check_xlrd():
    if not HAVE_XLRD:
        raise RuntimeError(
            "无法读取 .xls 文件：缺少 xlrd 库，请运行 pip install xlrd 后重试"
        )


def _xls_to_openpyxl(xls_path):
    """读取旧版 Excel 97-2003 (.xls, BIFF/OLE2) 文件，转为内存中的 openpyxl.Workbook。

    openpyxl 自身不支持 .xls，故用 xlrd 读取后重建为统一结构，
    供后续所有功能模块（合并/拆分/转 CSV）无差别处理。
    日期单元格转为 datetime，布尔单元格转为 bool，错误单元格置空。
    """
    _check_xlrd()
    try:
        book = xlrd.open_workbook(xls_path)
    except Exception as ex:
        raise RuntimeError(
            f"无法读取 .xls 文件（可能不是有效的 Excel 97-2003 文件或已损坏）: {ex}"
        )
    try:
        datemode = book.datemode
        out_wb = openpyxl.Workbook()
        out_wb.remove(out_wb.active)
        used = set()
        for idx in range(book.nsheets):
            src = book.sheet_by_index(idx)
            base = (src.name or f"Sheet{idx + 1}")[:31]
            name = base
            n = 1
            while name in used:
                suf = f"_{n}"
                name = base[:31 - len(suf)] + suf
                n += 1
            used.add(name)
            ws = out_wb.create_sheet(title=name)
            for r in range(src.nrows):
                row = []
                for c in range(src.ncols):
                    cell = src.cell(r, c)
                    val = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            val = xlrd.xldate.xldate_as_datetime(val, datemode)
                        except Exception:
                            # 日期序列号越界等异常：退回原始数值，避免整表读取失败
                            val = cell.value
                    elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                        val = bool(val)
                    elif cell.ctype == xlrd.XL_CELL_ERROR:
                        val = None
                    row.append(val)
                ws.append(row)
        return out_wb
    finally:
        try:
            book.release_resources()
        except Exception as ex:
            print(f"[warn] 释放 .xls 工作簿资源失败: {ex}", file=sys.stderr)


def load_workbook_any(path):
    """统一加载 Excel 文件，兼容 .xlsx/.xlsm（openpyxl）与 .xls（xlrd 转内存 Workbook）。

    返回对象支持 sheetnames / ws.iter_rows(values_only=True) / .close()，
    与 openpyxl.Workbook 接口一致；调用方无需关心底层格式。
    """
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        return _xls_to_openpyxl(path)
    # .xlsx / .xlsm 及未知扩展名交给 openpyxl（会给出清晰报错）
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as ex:
        msg = str(ex)
        if "zip" in msg.lower() or "not a zip" in msg.lower():
            raise RuntimeError(
                f"无法读取 Excel 文件「{os.path.basename(path)}」"
                f"（可能文件已损坏，或实为旧版 .xls 却以 .xlsx 命名）: {msg}"
            )
        raise RuntimeError(
            f"无法读取 Excel 文件「{os.path.basename(path)}」（可能已损坏或格式不受支持）: {msg}"
        )


class ExcelWorker:
    """Excel 工作簿核心逻辑"""

    @staticmethod
    def _check_openpyxl():
        if not HAVE_OPENPYXL:
            raise RuntimeError("需要 openpyxl 库来处理 Excel 文件，请先安装: pip install openpyxl")

    # ================================================================
    # F01 · 多工作表合并
    # ================================================================
    @staticmethod
    def merge_workbooks(file_paths, output_path, all_sheets=True,
                        progress_callback=None):
        """
        把多个 Excel 文件合并成一个工作簿。
        - all_sheets=True: 遍历每个文件的全部工作表
        - all_sheets=False: 仅合并每个文件的第一个工作表
        - 每个来源工作表作为独立工作表保留，命名: 文件名_工作表名
        - 不修改原文件
        """
        ExcelWorker._check_openpyxl()

        if not file_paths:
            return {"ok": False, "msg": "没有文件", "output": None}

        merged_wb = None
        try:
            merged_wb = openpyxl.Workbook()
            merged_wb.remove(merged_wb.active)

            total = len(file_paths)
            used_names = set()

            for idx, fp in enumerate(file_paths):
                if not os.path.exists(fp):
                    continue
                wb = load_workbook_any(fp)
                try:
                    fname = Path(fp).stem
                    sheets = wb.sheetnames if all_sheets else wb.sheetnames[:1]
                    for sn in sheets:
                        ws = wb[sn]
                        base_name = f"{fname}_{sn}"[:31]
                        name = base_name
                        counter = 1
                        while name in used_names:
                            suffix = f"_{counter}"
                            name = base_name[:31 - len(suffix)] + suffix
                            counter += 1
                        used_names.add(name)

                        new_ws = merged_wb.create_sheet(title=name)
                        # 逐行流式写入（不 list() 全量载入）
                        for row in ws.iter_rows(values_only=True):
                            new_ws.append(row)
                finally:
                    wb.close()

                if progress_callback:
                    progress_callback((idx + 1) / total)

            if progress_callback:
                progress_callback(1.0)

            merged_wb.save(output_path)
            return {
                "ok": True,
                "msg": f"已合并 {total} 个文件",
                "output": output_path,
            }
        except Exception as ex:
            return {"ok": False, "msg": str(ex), "output": None}
        finally:
            if merged_wb is not None:
                with contextlib.suppress(Exception):
                    merged_wb.close()

    # ================================================================
    # F05 · 单工作簿内表合并到一表
    # ================================================================
    @staticmethod
    def merge_sheets_in_workbook(src_path, output_path=None,
                                 add_source_column=True,
                                 progress_callback=None):
        """
        将一个 Excel 工作簿内的多个工作表合并成一个工作表。
        - 假设每个工作表的表头相同（只保留一次表头），列数不一致时跳过并警告
        - add_source_column: 是否新增一列标识来源工作表名
        - 输出到新文件
        """
        ExcelWorker._check_openpyxl()

        if output_path is None:
            parent = os.path.dirname(src_path)
            stem = Path(src_path).stem
            output_path = os.path.join(parent, f"{stem}_合并.xlsx")

        wb = None
        out_wb = None
        try:
            wb = load_workbook_any(src_path)
            total_sheets = len(wb.sheetnames)

            out_wb = openpyxl.Workbook()
            out_ws = out_wb.active
            out_ws.title = "合并结果"

            header_written = False
            header_col_count = 0
            source_col_name = "来源工作表"
            warnings = []

            for idx, sn in enumerate(wb.sheetnames):
                ws = wb[sn]
                header = None
                data_iter = ws.iter_rows(values_only=True)
                try:
                    header = next(data_iter)
                except StopIteration:
                    continue

                col_count = len(header) if header else 0
                if not header_written:
                    original_header = list(header) if header else []
                    header_col_count = col_count
                    if add_source_column:
                        original_header.append(source_col_name)
                    out_ws.append(original_header)
                    header_written = True
                else:
                    if col_count != header_col_count:
                        warnings.append(f"工作表「{sn}」列数({col_count})与首表({header_col_count})不一致，已跳过")
                        continue

                # 逐行流式写入（data_iter 在两个分支都指向同一迭代器，此处统一引用）
                for row in data_iter:
                    row_list = list(row) if row else []
                    if add_source_column:
                        row_list.append(sn)
                    out_ws.append(row_list)

                if progress_callback:
                    progress_callback((idx + 1) / total_sheets)

            wb.close()
            wb = None

            if progress_callback:
                progress_callback(1.0)

            out_wb.save(output_path)
            msg = f"已合并 {total_sheets} 个工作表"
            if warnings:
                msg += "（注意: " + "; ".join(warnings) + "）"
            return {
                "ok": True,
                "msg": msg,
                "output": output_path,
            }
        except Exception as ex:
            return {"ok": False, "msg": str(ex), "output": None}
        finally:
            if wb is not None:
                with contextlib.suppress(Exception):
                    wb.close()
            if out_wb is not None:
                with contextlib.suppress(Exception):
                    out_wb.close()

    # ================================================================
    # F09 · 工作簿拆分为 CSV 文件
    # ================================================================
    @staticmethod
    def workbook_to_csvs(src_path, skip_empty=True,
                         progress_callback=None):
        """
        将一个 Excel 工作簿拆分为多个 CSV 文件（每个工作表一个 CSV）。
        - 输出命名: 原文件名_工作表名.csv
        - 跳过空工作表
        - 编码: UTF-8-BOM
        - 不修改原文件
        """
        ExcelWorker._check_openpyxl()

        parent = os.path.dirname(src_path)
        stem = Path(src_path).stem

        wb = None
        try:
            wb = load_workbook_any(src_path)
            total_sheets = len(wb.sheetnames)
            parts = []

            for idx, sn in enumerate(wb.sheetnames):
                ws = wb[sn]
                safe_sn = sn.replace("/", "_").replace("\\", "_").replace(":", "_")
                safe_sn = safe_sn.replace("*", "_").replace("?", "_").replace('"', "_").replace("<", "_").replace(">", "_").replace("|", "_")
                out_name = f"{stem}_{safe_sn}.csv"
                out_path = os.path.join(parent, out_name)

                row_count = 0
                with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
                    writer = csv.writer(f)
                    for row in ws.iter_rows(values_only=True):
                        clean_row = [_cell_to_str(c) for c in row]
                        writer.writerow(clean_row)
                        row_count += 1

                if skip_empty and row_count == 0:
                    with contextlib.suppress(OSError):
                        os.remove(out_path)
                else:
                    parts.append(out_path)

                if progress_callback:
                    progress_callback((idx + 1) / total_sheets)

            if progress_callback:
                progress_callback(1.0)

            return {
                "ok": True,
                "msg": f"已拆分为 {len(parts)} 个 CSV 文件",
                "parts": parts,
                "total_parts": len(parts),
            }
        except Exception as ex:
            return {"ok": False, "msg": str(ex), "parts": []}
        finally:
            if wb is not None:
                with contextlib.suppress(Exception):
                    wb.close()

    # ================================================================
    # F10 · 超 256 列工作表拆分为多个工作表
    # ================================================================
    @staticmethod
    def split_wide_worksheet(src_path, threshold=256,
                             progress_callback=None):
        """
        当 Excel 工作表列数超过 threshold 时，按列拆分为多个工作表。
        - 每个分片都复制表头行
        - 分片命名: 原表名_1, 原表名_2, ...
        - 输出到新工作簿
        - 不修改原文件
        """
        ExcelWorker._check_openpyxl()

        parent = os.path.dirname(src_path)
        stem = Path(src_path).stem
        output_path = os.path.join(parent, f"{stem}_拆分列.xlsx")

        wb = None
        out_wb = None
        try:
            wb = load_workbook_any(src_path)
            out_wb = openpyxl.Workbook()
            out_wb.remove(out_wb.active)

            total_sheets = len(wb.sheetnames)

            for idx, sn in enumerate(wb.sheetnames):
                ws = wb[sn]
                # 流式读取：先拿 header，再逐行处理
                ws_iter = ws.iter_rows(values_only=True)
                try:
                    header = next(ws_iter)
                except StopIteration:
                    continue
                if not header:
                    continue

                total_cols = len(header)

                if total_cols <= threshold:
                    safe_name = sn[:31] if len(sn) > 31 else sn
                    new_ws = out_wb.create_sheet(title=safe_name)
                    new_ws.append(list(header))
                    for row in ws_iter:
                        new_ws.append(list(row) if row else [])
                else:
                    num_parts = (total_cols + threshold - 1) // threshold
                    # 缓存 header 分片
                    header_parts = []
                    for part_idx in range(num_parts):
                        start = part_idx * threshold
                        end = min(start + threshold, total_cols)
                        header_parts.append((start, end, list(header[start:end])))

                    # 为每个分片创建 sheet 并写表头
                    part_sheets = []
                    for part_idx, (start, end, part_header) in enumerate(header_parts):
                        part_name = f"{sn}_p{part_idx+1}"[:31]
                        new_ws = out_wb.create_sheet(title=part_name)
                        new_ws.append(part_header)
                        part_sheets.append((start, end, new_ws))

                    # 逐行流式切片写入
                    for row in ws_iter:
                        row_list = list(row) if row else []
                        for start, end, new_ws in part_sheets:
                            # 切片天然越界安全（#4修复）
                            new_ws.append(row_list[start:end])

                if progress_callback:
                    progress_callback((idx + 1) / total_sheets)

            wb.close()
            wb = None

            if progress_callback:
                progress_callback(1.0)

            total_output_sheets = len(out_wb.sheetnames)
            out_wb.save(output_path)
            return {
                "ok": True,
                "msg": f"已拆分列为 {total_output_sheets} 个工作表",
                "output": output_path,
            }
        except Exception as ex:
            return {"ok": False, "msg": str(ex), "output": None}
        finally:
            if wb is not None:
                with contextlib.suppress(Exception):
                    wb.close()
            if out_wb is not None:
                with contextlib.suppress(Exception):
                    out_wb.close()
