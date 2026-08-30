# -*- coding: utf-8 -*-
"""Api 类端到端冒烟测试（v2.3.0 拆分后新增）。

背景：v2.3.0 将 app_webview.py 的 9 个 exec_* 业务方法拆到 core/api_business.py
（BusinessApiMixin），主文件由 1169 行降至 599 行。拆分属高风险重构，
单元测试无法覆盖「Mixin 组合后方法是否仍可用」，故补充本集成测试：
**实例化真实 Api 类，跑通全部 9 个业务功能，校验输出文件真实生成且内容正确。**

实测价值：本项目历史上该手法两轮均抓到真实 bug（见 docs/审查记录/2026-08-28.md）。
"""
import os
import sys
import json
import time
import shutil
import tempfile
import unittest

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(ROOT, "src"))

from app_webview import Api  # noqa: E402

# 等待异步工作线程完成的超时（秒）。exec_* 均启动 daemon 线程，
# 完成后由 _on_done → _set_busy(False) 释放。
TIMEOUT = 30.0


def _wait_idle(api, timeout=TIMEOUT):
    """轮询等待 is_busy 归 False。返回 True=正常完成，False=超时。"""
    deadline = time.perf_counter() + timeout
    while time.perf_counter() < deadline:
        if not api.is_busy:
            return True
        time.sleep(0.05)
    return False


class TestApiIntegration(unittest.TestCase):
    """端到端：真实 Api 实例 + 真实文件 IO。"""

    @classmethod
    def setUpClass(cls):
        cls.tmp = tempfile.mkdtemp(prefix="niuniu_smoke_")
        # --- CSV 测试数据：中文 / 逗号 / 引号 / 日期列 / 多编码场景 ---
        cls.csv_hdr = "姓名,城市,金额,日期,备注"
        cls.csv_rows = [
            '张三,"上海,浦东",1200,2026-01-15,"含,逗号"',
            '李四,北京,850,2026-01-20,"含""引号"""',
            '王五,广州,2300,2026-02-03,普通备注',
            '赵六,深圳,450,2026-02-18,含中文逗号，测试',
            '钱七,杭州,1680,2026-03-05,末尾',
        ]
        cls.csv1 = os.path.join(cls.tmp, "销售数据A.csv")
        cls.csv2 = os.path.join(cls.tmp, "销售数据B.csv")
        # UTF-8-BOM：验证 BOM 检测与剥离
        with open(cls.csv1, "w", encoding="utf-8-sig", newline="") as f:
            f.write(cls.csv_hdr + "\n" + "\n".join(cls.csv_rows[:3]) + "\n")
        with open(cls.csv2, "w", encoding="utf-8-sig", newline="") as f:
            f.write(cls.csv_hdr + "\n" + "\n".join(cls.csv_rows[3:]) + "\n")

        # --- Excel 测试数据：两个工作簿，各 2 个 sheet ---
        try:
            from openpyxl import Workbook
            cls.xlsx1 = os.path.join(cls.tmp, "报表一.xlsx")
            cls.xlsx2 = os.path.join(cls.tmp, "报表二.xlsx")
            for path, tag in ((cls.xlsx1, "A"), (cls.xlsx2, "B")):
                wb = Workbook()
                ws = wb.active
                ws.title = f"主表{tag}"
                ws.append(["编号", "项目", "金额", "日期"])
                for i in range(1, 6):
                    ws.append([i, f"项目{tag}{i}", i * 100, f"2026-0{i}-01"])
                ws2 = wb.create_sheet(f"明细{tag}")
                ws2.append(["编号", "说明"])
                for i in range(1, 4):
                    ws2.append([i, f"明细{tag}{i}"])
                wb.save(path)
            cls.has_openpyxl = True
        except ImportError:
            cls.has_openpyxl = False

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmp, ignore_errors=True)

    def setUp(self):
        self.api = Api()
        # 无窗口时 UI 回调均有 `if self._window` 保护，此处保持 None 即可
        self.api._window = None

    # ---------- 辅助 ----------
    def _load(self, *paths):
        """把文件加入 Api.file_list，返回是否全部成功。"""
        res = self.api.add_files(json.dumps(list(paths)))
        self.assertTrue(res["ok"], f"add_files 失败: {res}")
        self.assertEqual(len(self.api.file_list), len(paths),
                         "file_list 数量与传入路径数不符")
        return res

    def _snapshot(self):
        """快照 tmp 目录当前全部文件（用于精确计算「本次新增」）。"""
        s = set()
        for dirpath, _dirs, files in os.walk(self.tmp):
            for f in files:
                s.add(os.path.join(dirpath, f))
        return s

    def _run(self, method, params):
        """调用 exec_* 并等待完成，返回 (msg, 本次**新增**的文件列表)。

        采用执行前后快照取差集，而非「统计 out_dir 下所有文件」——
        因为 out_dir 模式下部分功能会先把源文件复制进输出目录
        （如 _split_rows_worker 的 shutil.copy2），直接统计会混入源副本。
        """
        before = self._snapshot()
        fn = getattr(self.api, method)
        res = fn(json.dumps(params))
        self.assertTrue(res.get("ok"), f"{method} 返回失败: {res}")
        self.assertTrue(_wait_idle(self.api), f"{method} 超时未完成（>{TIMEOUT}s）")
        after = self._snapshot()
        return res.get("msg", ""), sorted(after - before)

    def _out_dir(self, tag):
        d = os.path.join(self.tmp, f"out_{tag}")
        os.makedirs(d, exist_ok=True)
        return d

    # ---------- 1. 编码转换 ----------
    def test_01_exec_convert(self):
        out = self._out_dir("convert")
        self._load(self.csv1)
        _, files = self._run("exec_convert",
                             {"encoding": "utf-8", "copy_mode": True,
                              "out_dir": out, "open_folder": False})
        produced = [f for f in files if f.startswith(out)]
        self.assertTrue(produced, f"转换未生成输出文件，目录内容: {os.listdir(out)}")
        # 校验内容完整：表头 + 3 行数据，且逗号/引号字段未被破坏
        with open(produced[0], encoding="utf-8") as f:
            lines = [ln for ln in f.read().splitlines() if ln.strip()]
        self.assertEqual(len(lines), 4, f"行数不符（应为表头+3行）: {lines}")
        self.assertIn("上海,浦东", lines[1], "含逗号的字段被破坏（CSV 序列化有问题）")

    # ---------- 2. 按行分割 ----------
    def test_02_exec_split_rows(self):
        """max_rows 语义（易误解，务必保持注释与实现同步）：
        max_rows = **每个输出文件的总行数上限（含表头）**，与 UI 文案
        「每个文件行数」(app.html:433) 一致。
        实现见 core/split.py:62 —— per_part = max_rows - 1（扣掉表头），
        且前置判断 `total <= max_rows` 的 total 同样含表头（count_lines）。
        故 max_rows=2、数据 3 行 → 每片 1 表头+1 数据 = 2 行 → 3 片。
        """
        out = self._out_dir("split_rows")
        self._load(self.csv1)
        _, new_files = self._run("exec_split_rows",
                                 {"max_rows": 2, "out_dir": out,
                                  "encoding": "utf-8", "open_folder": False})
        # out_dir 模式会先把源文件复制进输出目录，故新增 = 1 源副本 + N 个分片
        parts = sorted(f for f in new_files if "_part" in os.path.basename(f))
        self.assertEqual(len(parts), 3,
                         f"分片数不符（max_rows=2 → 每片2行含表头 → 3片）: {parts}")
        # 每片行数不得超过 max_rows，且首行必须是表头
        total_data = 0
        for p in parts:
            with open(p, encoding="utf-8-sig") as f:
                lines = [ln for ln in f.read().splitlines() if ln.strip()]
            self.assertLessEqual(len(lines), 2, f"分片超过 max_rows: {p}")
            self.assertEqual(lines[0].strip(), self.csv_hdr, f"分片缺少表头: {p}")
            total_data += len(lines) - 1
        self.assertEqual(total_data, 3, f"分片后数据行总量丢失: {total_data} != 3")

    # ---------- 3. 按日期分割 ----------
    def test_03_exec_split_date(self):
        out = self._out_dir("split_date")
        self._load(self.csv1)
        _, files = self._run("exec_split_date",
                             {"date_col": "日期", "granularity": "month",
                              "out_dir": out, "encoding": "utf-8",
                              "open_folder": False})
        produced = [f for f in files if f.startswith(out)]
        self.assertTrue(produced, f"按日期分割未生成文件: {os.listdir(out)}")

    # ---------- 4. CSV 合并 ----------
    def test_04_exec_csv_merge(self):
        out = self._out_dir("csv_merge")
        self._load(self.csv1, self.csv2)
        _, files = self._run("exec_csv_merge",
                             {"mode": "merge", "output_name": "合并结果.csv",
                              "out_dir": out, "sort": "", "open_folder": False})
        produced = [f for f in files if f.startswith(out)]
        self.assertTrue(produced, f"CSV 合并未生成文件: {os.listdir(out)}")
        with open(produced[0], encoding="utf-8") as f:
            lines = [ln for ln in f.read().splitlines() if ln.strip()]
        # 两文件各 3/2 行数据，合并后数据行应为 5（表头不重复）
        self.assertEqual(len(lines), 6, f"合并行数不符（表头1+数据5）: {lines}")

    # ---------- 5. 关键词提取 ----------
    def test_05_exec_keyword(self):
        out = self._out_dir("keyword")
        self._load(self.csv1)
        _, files = self._run("exec_keyword",
                             {"keyword": "上海", "match_mode": "contains",
                              "col_name": "", "out_dir": out,
                              "open_folder": False})
        produced = [f for f in files if f.startswith(out)]
        self.assertTrue(produced, f"关键词提取未生成文件: {os.listdir(out)}")
        with open(produced[0], encoding="utf-8") as f:
            content = f.read()
        self.assertIn("张三", content, "关键词命中行未提取到")
        self.assertNotIn("李四", content, "非命中行不应出现在结果中")

    @unittest.skipUnless(True, "")
    def test_06_exec_wb_merge(self):
        if not self.has_openpyxl:
            self.skipTest("openpyxl 未安装")
        out = self._out_dir("wb_merge")
        self._load(self.xlsx1, self.xlsx2)
        _, files = self._run("exec_wb_merge",
                             {"mode": "merge", "output_name": "工作簿合并.xlsx",
                              "out_dir": out, "open_folder": False})
        produced = [f for f in files if f.startswith(out) and f.endswith(".xlsx")]
        self.assertTrue(produced, f"工作簿合并未生成文件: {os.listdir(out)}")
        # 校验 sheet 数：两个工作簿各 2 sheet → 合并后应 ≥ 2
        from openpyxl import load_workbook
        wb = load_workbook(produced[0])
        self.assertGreaterEqual(len(wb.sheetnames), 2,
                                f"合并后 sheet 数异常: {wb.sheetnames}")
        wb.close()

    def test_07_exec_sheet_merge(self):
        if not self.has_openpyxl:
            self.skipTest("openpyxl 未安装")
        out = self._out_dir("sheet_merge")
        self._load(self.xlsx1)
        _, files = self._run("exec_sheet_merge",
                             {"header_strategy": "first", "out_dir": out,
                              "open_folder": False})
        produced = [f for f in files if f.startswith(out) and f.endswith(".xlsx")]
        self.assertTrue(produced, f"工作表合并未生成文件: {os.listdir(out)}")

    def test_08_exec_wb_split(self):
        if not self.has_openpyxl:
            self.skipTest("openpyxl 未安装")
        out = self._out_dir("wb_split")
        self._load(self.xlsx1)
        _, files = self._run("exec_wb_split",
                             {"out_dir": out, "encoding": "utf-8",
                              "open_folder": False})
        produced = [f for f in files if f.startswith(out)]
        self.assertTrue(produced, f"工作簿拆分未生成文件: {os.listdir(out)}")

    def test_09_exec_wide_split(self):
        if not self.has_openpyxl:
            self.skipTest("openpyxl 未安装")
        out = self._out_dir("wide_split")
        self._load(self.xlsx1)
        _, files = self._run("exec_wide_split",
                             {"cols": "项目", "out_dir": out,
                              "open_folder": False})
        produced = [f for f in files if f.startswith(out)]
        self.assertTrue(produced, f"宽表拆分未生成文件: {os.listdir(out)}")

    # ---------- 契约完整性：9 个 exec_* 必须全部可调用 ----------
    def test_10_all_exec_methods_callable(self):
        expected = [
            "exec_convert", "exec_split_rows", "exec_split_date",
            "exec_csv_merge", "exec_wb_merge", "exec_sheet_merge",
            "exec_keyword", "exec_wb_split", "exec_wide_split",
        ]
        for m in expected:
            self.assertTrue(hasattr(self.api, m), f"Api 缺少方法 {m}（拆分后契约断裂）")
            self.assertTrue(callable(getattr(self.api, m)), f"{m} 不可调用")

    # ---------- 窗口控制方法仍在主类（未随业务拆分丢失） ----------
    def test_11_window_methods_present(self):
        for m in ("minimize_window", "toggle_maximize", "close_window",
                  "resize_step", "move_step", "get_version"):
            self.assertTrue(hasattr(self.api, m), f"窗口方法 {m} 丢失")

    def test_12_get_version_matches_VERSION_file(self):
        with open(os.path.join(ROOT, "VERSION"), encoding="utf-8") as f:
            expect = f.read().strip()
        self.assertEqual(self.api.get_version()["version"], expect,
                         "get_version 与 VERSION 文件不一致")


if __name__ == "__main__":
    unittest.main(verbosity=2)
