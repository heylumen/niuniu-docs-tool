#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""牛牛文档工具 — 后端核心功能无头功能测试

覆盖 9 个功能模块的完整业务流程与边界场景，不依赖 Tk 图形界面，
可直接在无显示环境（如打包沙箱）运行，验证 core/ 业务逻辑无回归。

运行:
    python -m unittest tests.test_core_functional -v
或经提交门禁 scripts/review_check.py 自动发现执行。
"""
import csv
import os
import sys
import tempfile
import unittest

_src = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "src")
if _src not in sys.path:
    sys.path.insert(0, _src)

from core.encoding import EncodingDetector
from core.convert import FileConverter
from core.split import CsvSplitter
from core.merge import CsvMerger
from core.extract import CsvExtractor
from core.sheet import ExcelWorker, HAVE_OPENPYXL


def _write_text(path, text, encoding="utf-8"):
    with open(path, "w", encoding=encoding, newline="") as f:
        f.write(text)


def _read_csv_rows(path):
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return list(csv.reader(f))


def _has_bom(path):
    with open(path, "rb") as f:
        return f.read(3) == b"\xef\xbb\xbf"


class TestConvert(unittest.TestCase):
    """F01 编码转换：GBK→UTF-8-BOM / UTF-8 补 BOM / 未知编码拒绝"""

    def test_gbk_to_utf8_bom(self):
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "gbk.csv")
            _write_text(src, "姓名,金额\n张三,100\n李四,200\n", encoding="gbk")
            dst = os.path.join(d, "out.csv")
            r = FileConverter.convert(src, dst)
            self.assertTrue(r["ok"], r)
            self.assertTrue(_has_bom(dst), "输出应带 UTF-8 BOM")
            rows = _read_csv_rows(dst)
            self.assertEqual(rows[0], ["姓名", "金额"])
            self.assertEqual(rows[1], ["张三", "100"])

    def test_utf8_no_bom_added(self):
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "u8.csv")
            _write_text(src, "a,b\n1,2\n", encoding="utf-8")  # 无 BOM
            dst = os.path.join(d, "out.csv")
            r = FileConverter.convert(src, dst)
            self.assertTrue(r["ok"])
            self.assertTrue(_has_bom(dst))

    def test_already_bom_copied(self):
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "bom.csv")
            _write_text(src, "a,b\n1,2\n", encoding="utf-8-sig")
            dst = os.path.join(d, "out.csv")
            r = FileConverter.convert(src, dst)
            self.assertTrue(r["ok"])
            self.assertIn("直接复制", r["msg"])

    def test_unknown_encoding_rejected(self):
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "bin.bin")
            with open(src, "wb") as f:
                f.write(b"\xff\xfe\x00\x01\x02")  # 非文本
            self.assertEqual(EncodingDetector.detect(src)[0], "unknown")
            dst = os.path.join(d, "out.csv")
            r = FileConverter.convert(src, dst)
            self.assertFalse(r["ok"])
            self.assertIn("未知编码", r["msg"])


class TestSplitRows(unittest.TestCase):
    """split_rows：分片数量 / 每片含表头 / 边界（≤阈值跳过）"""

    def _make(self, d, n):
        p = os.path.join(d, "big.csv")
        lines = ["id,val"] + [f"{i},{i*2}" for i in range(n)]
        _write_text(p, "\n".join(lines) + "\n", encoding="utf-8")
        return p

    def test_split_into_parts(self):
        with tempfile.TemporaryDirectory() as d:
            p = self._make(d, 2500)
            r = CsvSplitter.split_by_rows(p, max_rows=1000)
            self.assertTrue(r["ok"])
            self.assertEqual(len(r["parts"]), 3, r)
            for part in r["parts"]:
                rows = _read_csv_rows(part)
                self.assertEqual(rows[0], ["id", "val"], "每片必须含表头")

    def test_boundary_no_split(self):
        with tempfile.TemporaryDirectory() as d:
            # 500 数据行 + 1 表头 = 501 物理行 ≤ 阈值 1000 → 应跳过不分割
            p = self._make(d, 500)
            r = CsvSplitter.split_by_rows(p, max_rows=1000)
            self.assertTrue(r["ok"])
            self.assertTrue(r.get("skipped"))
            self.assertEqual(len(r["parts"]), 0)


class TestSplitDate(unittest.TestCase):
    """split_date：按月份拆分 / 未知日期归入单独文件"""

    def test_split_by_month(self):
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "tx.csv")
            header = "交易日期,金额\n"
            rows = ["2026-01-15,10", "2026-01-20,20", "2026-02-03,30"]
            _write_text(p, header + "\n".join(rows) + "\n", encoding="utf-8")
            r = CsvSplitter.split_by_date(p, date_column="交易日期", granularity="month")
            self.assertTrue(r["ok"], r)
            self.assertEqual(len(r["parts"]), 2, "应拆成 2 个月")
            names = [os.path.basename(x) for x in r["parts"]]
            self.assertTrue(any("2026-01" in n for n in names))
            self.assertTrue(any("2026-02" in n for n in names))

    def test_unknown_date_bucket(self):
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "tx.csv")
            _write_text(p, "交易日期,金额\n无日期,99\n", encoding="utf-8")
            r = CsvSplitter.split_by_date(p, date_column="交易日期", granularity="day")
            self.assertTrue(r["ok"])
            self.assertTrue(any("未知日期" in os.path.basename(x) for x in r["parts"]))


class TestCsvMerge(unittest.TestCase):
    """csv_merge：快速拷贝 / 删表头 / 加文件名列"""

    def _two(self, d):
        a = os.path.join(d, "a.csv")
        b = os.path.join(d, "b.csv")
        _write_text(a, "x,y\n1,2\n", encoding="utf-8")
        _write_text(b, "x,y\n3,4\n", encoding="utf-8")
        return a, b

    def test_merge_fast_single_header(self):
        with tempfile.TemporaryDirectory() as d:
            a, b = self._two(d)
            out = os.path.join(d, "m.csv")
            r = CsvMerger.merge_fast([a, b], out)
            self.assertTrue(r["ok"], r)
            rows = _read_csv_rows(out)
            self.assertEqual(rows[0], ["x", "y"])
            self.assertEqual(rows[1:], [["1", "2"], ["3", "4"]])

    def test_merge_remove_headers(self):
        with tempfile.TemporaryDirectory() as d:
            a, b = self._two(d)
            out = os.path.join(d, "m.csv")
            r = CsvMerger.merge_remove_headers([a, b], out)
            self.assertTrue(r["ok"])
            rows = _read_csv_rows(out)
            self.assertEqual(rows.count(["x", "y"]), 1, "表头只应出现一次")

    def test_merge_with_filename(self):
        with tempfile.TemporaryDirectory() as d:
            a, b = self._two(d)
            out = os.path.join(d, "m.csv")
            r = CsvMerger.merge_with_filename([a, b], out)
            self.assertTrue(r["ok"])
            rows = _read_csv_rows(out)
            self.assertEqual(rows[0][-1], "来源文件")
            self.assertIn("a.csv", rows[1])


class TestKeyword(unittest.TestCase):
    """keyword：包含匹配仅输出匹配行；边界：空关键字拒绝"""

    def test_contains_match(self):
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "k.csv")
            _write_text(p, "备注,金额\n退款100,1\n正常200,2\n退款50,3\n", encoding="utf-8")
            out = os.path.join(d, "r.csv")
            r = CsvExtractor.extract_by_keyword(p, ["退款"], output_path=out,
                                                match_mode="contains")
            self.assertTrue(r["ok"], r)
            self.assertEqual(r["matched_count"], 2)
            rows = _read_csv_rows(out)
            self.assertEqual(len(rows), 3)  # 表头 + 2 匹配

    def test_empty_keywords(self):
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "k.csv")
            _write_text(p, "a,b\n1,2\n", encoding="utf-8")
            r = CsvExtractor.extract_by_keyword(p, [], output_path=os.path.join(d, "r.csv"))
            self.assertFalse(r["ok"])


@unittest.skipUnless(HAVE_OPENPYXL, "需要 openpyxl 才能测试 Excel 功能")
class TestExcel(unittest.TestCase):
    """wb_merge / sheet_merge / wb_split / wide_split 四个 Excel 后端"""

    def _wb(self, path, sheets):
        import openpyxl
        wb = openpyxl.Workbook()
        for i, (name, data) in enumerate(sheets):
            ws = wb.active if i == 0 else wb.create_sheet()
            ws.title = name
            for row in data:
                ws.append(row)
        wb.save(path)

    def test_merge_workbooks(self):
        with tempfile.TemporaryDirectory() as d:
            a = os.path.join(d, "a.xlsx")
            b = os.path.join(d, "b.xlsx")
            self._wb(a, [("S1", [["h1", "h2"], [1, 2]])])
            self._wb(b, [("T1", [["h1", "h2"], [3, 4]])])
            out = os.path.join(d, "merged.xlsx")
            r = ExcelWorker.merge_workbooks([a, b], out)
            self.assertTrue(r["ok"], r)
            import openpyxl
            mwb = openpyxl.load_workbook(out)
            names = mwb.sheetnames
            self.assertEqual(len(names), 2)
            self.assertTrue(any("a_S1" in n for n in names))
            self.assertTrue(any("b_T1" in n for n in names))

    def test_merge_sheets_in_workbook(self):
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "s.xlsx")
            self._wb(src, [("S1", [["h1", "h2"], [1, 2]]),
                            ("S2", [["h1", "h2"], [3, 4]])])
            out = os.path.join(d, "sm.xlsx")
            r = ExcelWorker.merge_sheets_in_workbook(src, out)
            self.assertTrue(r["ok"], r)
            import openpyxl
            wb = openpyxl.load_workbook(out)
            ws = wb["合并结果"]
            rows = list(ws.iter_rows(values_only=True))
            self.assertEqual(rows[0][-1], "来源工作表", "应有来源工作表列")
            self.assertEqual(len(rows), 3)  # 一次表头 + 2 数据行

    def test_workbook_to_csvs(self):
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "w.xlsx")
            self._wb(src, [("S1", [["a", "b"], [1, 2]]),
                            ("S2", [["c", "d"], [3, 4]])])
            r = ExcelWorker.workbook_to_csvs(src)
            self.assertTrue(r["ok"], r)
            self.assertEqual(len(r["parts"]), 2, "每个工作表一个 CSV")
            for p in r["parts"]:
                self.assertTrue(os.path.exists(p) and p.endswith(".csv"))

    def test_split_wide_worksheet(self):
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "wide.xlsx")
            header = [f"c{i}" for i in range(300)]
            data = list(range(300))
            self._wb(src, [("W", [header, data])])
            r = ExcelWorker.split_wide_worksheet(src, threshold=256)
            self.assertTrue(r["ok"], r)
            import openpyxl
            wb = openpyxl.load_workbook(r["output"])
            self.assertGreaterEqual(len(wb.sheetnames), 2, "300 列应拆为 ≥2 片")


if __name__ == "__main__":
    unittest.main(verbosity=2)
